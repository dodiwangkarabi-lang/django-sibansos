import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font

from reportlab.platypus import SimpleDocTemplate, Image, Table, TableStyle, Spacer, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

from django.conf import settings

# models
from penyaluran.models import (
    PenerimaBantuan
)



def format_header(col_name: str) -> str:
    return col_name.replace("_", " ").upper()

def pengajuan_to_table(queryset):
    data = [[
        "No",
        "NIK",
        "Nama",
        "Alamat",
        "No HP",
        "Jenis Bantuan",
        "Tanggal",
        "Status"
    ]]
    
    styles = getSampleStyleSheet()
    def p(text):
        return Paragraph(str(text), styles["BodyText"])

    no = 1
    for obj in queryset:
        data.append([
            no,
            p(obj.masyarakat.nik),
            p(obj.masyarakat.nama),
            p(obj.masyarakat.alamat),
            p(obj.masyarakat.no_hp),
            p(obj.bantuan.nama),
            p(obj.tanggal_terima.strftime("%d-%m-%Y")),
            "Diterima",
        ])
        no += 1

    table = Table(
        data, hAlign="LEFT",
        colWidths=[
            30,   # No
            70,   # NIK
            80,  # Nama
            80,  # Alamat
            80,   # No HP
            80,  # Jenis Bantuan
            65,   # Tanggal
            50,   # Status
        ],
    )

    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    return table

def dataframe_to_table(df, **kwargs):
    data = [df.columns.tolist()] + df.values.tolist()

    table = Table(data, hAlign="LEFT")

    table.setStyle(TableStyle([
        # garis tabel saja
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

        # alignment
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        # header dibuat tebal saja (tanpa background)
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        # padding biar tidak rapat
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
    ]))

    return table

def queryset_to_table(queryset, fields, headers=None, **kwargs):
    """
    queryset: Django QuerySet
    fields: list field yang mau ditampilkan, contoh ["nama", "nilai"]
    headers: optional custom header
    """

    # header
    if headers is None:
        headers = [f.capitalize() for f in fields]

    data = [headers]

    # isi data dari queryset
    for obj in queryset:
        row = [getattr(obj, field) for field in fields]
        data.append(row)

    table = Table(data, hAlign="LEFT")

    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
    ]))

    return table

def build_pdf(df, **kwargs):
    """
    bangun pdf

    Args:
        df (): DataFrame atau QuerySet

    Returns:
        _type_: _description_
    """
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []

    # 1. HEADER IMAGE
    header_img = settings.MEDIA_ROOT / "lainnya" / "header.png"
    header = Image(header_img, width=500, height=80)
    elements.append(header)
    elements.append(Spacer(1, 20))
    
    # tambahan
    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("MyTitle")
    title_style.fontSize = 10
    title_style.leading = 14

    # Judul
    title = Paragraph(
        "<b>LAPORAN PENERIMA BANTUAN</b>",
        title_style
    )
    
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informasi bantuan
    info_data = [
        ["NAMA PROGRAM", ":", "Program Bantuan Sosial Masyarakat"],
        ["KATEGORI BANTUAN", ":", "Bantuan Tunai"],
        ["SUMBER DANA", ":", "APBD Kabupaten"],
        ["TAHUN", ":", "2026"],
        ["DISTRIK/KAMPUNG", ":", "Distrik Anggi / Kampung Testega"],
    ]

    info_table = Table(
        info_data,
        colWidths=[170, 15, 300],
        hAlign="LEFT"
    )

    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    elements.append(info_table)
    elements.append(Spacer(1, 20))

    # 2. TABLE DARI DATAFRAME
    # table = dataframe_to_table(df)
    table = pengajuan_to_table(df)
    # table = queryset_to_table(
    #     df,
    #     fields=["nik", "nama", "alamat", "no_hp", "jenis_bantuan", "tanggal_terima"],
    #     headers=["NIK", "NAMA", "ALAMAT", "NO HP", "JENIS BANTUAN", "TANGGAL TERIMA"]
    # )
    elements.append(table)
    elements.append(Spacer(1, 40))
    
    

    # 3. TANDA TANGAN DI AKHIR
    ttd_img = settings.MEDIA_ROOT / "lainnya" / "ttd.png"
    signature = Image(ttd_img, width=150, height=80)
    # sisipkan ke sebelah kanan
    signature.hAlign = "RIGHT"

    elements.append(signature)

    doc.build(elements)
    
    buffer.seek(0)
    return buffer

class Laporan:
    def __init__(self):
        pass
    
    def laporan_file(self, dataframe) -> io.BytesIO:
        """
        Return
            buffer
        """
        pdf = build_pdf(dataframe)
        return pdf
    
    def _generate_dataframe(self) -> pd.DataFrame:
        penerima_bantuan_qs = PenerimaBantuan.objects.all()
        no = 1
        penerima_bantuan_list = [
            {
                "nik": penerima_bantuan.masyarakat.nik,
                "nama": penerima_bantuan.masyarakat.nama,
                "alamat": penerima_bantuan.masyarakat.alamat,
                "no_hp": penerima_bantuan.masyarakat.no_hp,
                "jenis_bantuan": penerima_bantuan.bantuan.nama,
                "tanggal_terima": penerima_bantuan.tanggal_terima,
            }
            for penerima_bantuan in penerima_bantuan_qs
        ]
        penerima_bantuan_df = pd.DataFrame(penerima_bantuan_list)
        penerima_bantuan_df["no"] = penerima_bantuan_df.index + 1
        # pindahkan kolom no di awal
        penerima_bantuan_df = penerima_bantuan_df[["no"] + [col for col in penerima_bantuan_df.columns if col != "no"]]
        
        # rubah format tanggal menjadi tanggal-bulan-tahun
        penerima_bantuan_df["tanggal_terima"] = pd.to_datetime(penerima_bantuan_df["tanggal_terima"])
        penerima_bantuan_df["tanggal_terima"] = penerima_bantuan_df["tanggal_terima"].dt.strftime("%d-%m-%Y")
        
        return penerima_bantuan_df
    
    def _data_keterangan(self):
        keterangan = {
            "nama_program": "Bantuan Sosial Keluarga Sejahtera",
            "kategori_bantuan": "Sembako",
            "sumber_dana": "APBN",
            "tahun": "2026",
        }
        
        return keterangan
    
    
    def laporan_excel(self, nama_template=None, lokasi_simpan="laporan_final.xlsx") -> io.BytesIO:
        keterangan = self._data_keterangan()
        df = self._generate_dataframe()
        
        # =========================
        # LOAD TEMPLATE
        # =========================
        if nama_template is None:
            nama_template = settings.MEDIA_ROOT / "lainnya" / "template_excel.xlsx"
        wb = load_workbook(nama_template)
        
        ws = wb.active

        # =========================
        # ISI KETERANGAN
        # =========================
        ws["D10"] = keterangan["nama_program"]
        ws["D11"] = keterangan["kategori_bantuan"]
        ws["D12"] = keterangan["sumber_dana"]
        ws["D13"] = keterangan["tahun"]

        # =========================
        # TEMPAT TABEL START
        # =========================
        start_row = 16
        start_col = 2
        
        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        # header
        for col_idx, col_name in enumerate(df.columns, start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=format_header(col_name))
            cell.border = border
            cell.font = Font(bold=True, size=12, name="Times New Roman")
                
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.font = Font(bold=False, size=12, name="Times New Roman")
                
        # hitung posisi akhir data
        last_row = start_row + len(df)

        # posisi tanda tangan
        ttd_row = last_row + 3
        
        path_gambar = settings.MEDIA_ROOT / "lainnya" / "ttd_laporan.png"
        img = Image(path_gambar)
        ws.add_image(img, f"F{ttd_row}")

        # ws[f"D{ttd_row}"] = "Tanda Tangan"
        # ws[f"D{ttd_row + 3}"] = "Nama Penandatangan"

        # =========================
        # SAVE FILE BARU
        # =========================
        # wb.save(lokasi_simpan)
        # print("Sukses generate laporan dari template")
        
        # output
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def laporan_pdf(self, data):
        return